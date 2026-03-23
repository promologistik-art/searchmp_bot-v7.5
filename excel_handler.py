import io
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict


def _autofit_column_widths(worksheet):
    """Выравнивает ширину столбцов по содержимому."""
    for col_idx, column_cells in enumerate(worksheet.iter_cols(min_row=1, max_row=worksheet.max_row), 1):
        max_len = 10
        for cell in column_cells:
            try:
                val = cell.value
                if val is not None:
                    length = min(len(str(val)) + 1, 60)
                    max_len = max(max_len, length)
            except (TypeError, ValueError):
                pass
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_len


def _apply_header_style(worksheet):
    fill = PatternFill("solid", fgColor="FFEFEFEF")
    font = Font(bold=True)
    alignment = Alignment(vertical="center", wrap_text=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def _apply_user_input_column_style(worksheet, header_names: list[str]):
    # Excel-like light green fill for input columns
    green_fill = PatternFill("solid", fgColor="FFC6EFCE")
    alignment = Alignment(vertical="center")

    headers = [worksheet.cell(1, c).value for c in range(1, worksheet.max_column + 1)]
    for name in header_names:
        if name not in headers:
            continue
        col = headers.index(name) + 1
        # header cell
        hcell = worksheet.cell(1, col)
        hcell.fill = green_fill
        hcell.alignment = Alignment(vertical="center", wrap_text=True)
        # column cells
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row, col)
            cell.fill = green_fill
            cell.alignment = alignment


def _apply_fixed_column_widths_like_example(worksheet):
    # Ширины столбцов сняты с файла-примера ozon_11cats_20260312_104222.xlsx
    widths = {
        "A": 15.85546875,
        "B": 19.0,
        "C": 19.140625,
        "D": 11.0,
        "E": 19.0,
        "F": 14.0,
        "G": 13.0,
        "H": 15.0,
        "I": 10.0,
        "J": 13.0,
        "K": 12.140625,
        "L": 14.0,
        "M": 13.0,
        "N": 13.0,
        "O": 13.0,
        "P": 10.5703125,
        "Q": 8.0,
    }
    for col, width in widths.items():
        worksheet.column_dimensions[col].width = width


def create_excel_report(results: List[Dict]) -> io.BytesIO:
    """Создает Excel отчет с результатами анализа"""
    if not results:
        df = pd.DataFrame([{'Статус': 'Нет данных'}])
    else:
        df = pd.DataFrame(results)

        # Приводим входные ключи к новой структуре отчёта
        df = df.rename(columns={
            "category": "Категория",
            "name": "Название товара",
            "price": "Цена, р",
            "revenue": "Выручка за 30 дней",
            "competitors": "Количество конкурентов",
        })

        # Удаляем столбцы Бренд и Продавец (если присутствуют)
        for drop_col in ("brand", "seller"):
            if drop_col in df.columns:
                df = df.drop(columns=[drop_col])

        if "url" in df.columns:
            df["Ссылка на Ozon"] = df["url"]
            df = df.drop(columns=["url"])

        # Пользовательские/расчётные колонки
        df["Кол-во к закупке"] = ""
        df["Себестоимость"] = ""
        df["Комиссия"] = ""
        df["Логистика"] = ""
        df["Эквайринг"] = ""
        df["Всего расходы"] = ""
        df["Закуп итого, р"] = ""
        df["Прибыль на ед, р"] = ""
        df["Прибыль на партию, р"] = ""
        df["Маржа"] = ""
        df["ROI"] = ""

        # Итоговый порядок столбцов (как в твоём примере)
        col_order = [
            "Ссылка на Ozon",
            "Категория",
            "Название товара",
            "Цена, р",
            "Выручка за 30 дней",
            "Количество конкурентов",
            "Кол-во к закупке",
            "Себестоимость",
            "Комиссия",
            "Логистика",
            "Эквайринг",
            "Всего расходы",
            "Закуп итого, р",
            "Прибыль на ед, р",
            "Прибыль на партию, р",
            "Маржа",
            "ROI",
        ]
        df = df[[c for c in col_order if c in df.columns]]

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Результаты анализа')

        worksheet = writer.sheets['Результаты анализа']

        if results:
            # Индексы колонок по текущей шапке (на случай будущих перестановок)
            headers = [worksheet.cell(1, c).value for c in range(1, worksheet.max_column + 1)]
            def col_idx(name: str) -> int | None:
                try:
                    return headers.index(name) + 1
                except ValueError:
                    return None

            c_link = col_idx("Ссылка на Ozon")
            c_price = col_idx("Цена, р")
            c_qty = col_idx("Кол-во к закупке")
            c_cogs = col_idx("Себестоимость")
            c_fee = col_idx("Комиссия")
            c_log = col_idx("Логистика")
            c_acq = col_idx("Эквайринг")
            c_total = col_idx("Всего расходы")
            c_buy_total = col_idx("Закуп итого, р")
            c_profit_unit = col_idx("Прибыль на ед, р")
            c_profit_batch = col_idx("Прибыль на партию, р")
            c_margin = col_idx("Маржа")
            c_roi = col_idx("ROI")
            c_rev30 = col_idx("Выручка за 30 дней")

            for row in range(2, worksheet.max_row + 1):
                # Кликабельная ссылка
                if c_link is not None:
                    cell = worksheet.cell(row=row, column=c_link)
                    if isinstance(cell.value, str) and cell.value.startswith("http"):
                        # Для старых версий Excel надёжнее формула HYPERLINK()
                        url = cell.value.replace('"', '""')
                        cell.value = f'=HYPERLINK("{url}","{url}")'
                        cell.style = "Hyperlink"

                # Всего расходы (на единицу)
                if c_total is not None and None not in (c_cogs, c_fee, c_log, c_acq):
                    worksheet.cell(
                        row=row,
                        column=c_total,
                        value=f"={get_column_letter(c_cogs)}{row}"
                              f"+{get_column_letter(c_fee)}{row}"
                              f"+{get_column_letter(c_log)}{row}"
                              f"+{get_column_letter(c_acq)}{row}",
                    )

                # Закуп итого (на партию) = Всего расходы * Кол-во к закупу
                if c_buy_total is not None and None not in (c_total, c_qty):
                    worksheet.cell(
                        row=row,
                        column=c_buy_total,
                        value=f"={get_column_letter(c_total)}{row}*{get_column_letter(c_qty)}{row}",
                    )

                # Прибыль на единицу = Цена - Всего расходы
                if c_profit_unit is not None and None not in (c_price, c_total):
                    worksheet.cell(
                        row=row,
                        column=c_profit_unit,
                        value=f"={get_column_letter(c_price)}{row}-{get_column_letter(c_total)}{row}",
                    )

                # Прибыль на партию = Прибыль на ед * Кол-во к закупу
                if c_profit_batch is not None and None not in (c_profit_unit, c_qty):
                    worksheet.cell(
                        row=row,
                        column=c_profit_batch,
                        value=f"={get_column_letter(c_profit_unit)}{row}*{get_column_letter(c_qty)}{row}",
                    )

                # Маржа (%) = Прибыль на ед / Цена
                if c_margin is not None and None not in (c_profit_unit, c_price):
                    worksheet.cell(
                        row=row,
                        column=c_margin,
                        value=f"=IF({get_column_letter(c_price)}{row}>0,"
                              f"{get_column_letter(c_profit_unit)}{row}/{get_column_letter(c_price)}{row}*100,\"\")",
                    )

                # ROI (%) = Прибыль на ед / Всего расходы
                if c_roi is not None and None not in (c_profit_unit, c_total):
                    worksheet.cell(
                        row=row,
                        column=c_roi,
                        value=f"=IF({get_column_letter(c_total)}{row}>0,"
                              f"{get_column_letter(c_profit_unit)}{row}/{get_column_letter(c_total)}{row}*100,\"\")",
                    )

            # Форматы
            rub_fmt = '#,##0\\ _₽'
            pct_fmt = '0.00'
            if c_price is not None:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=c_price).number_format = rub_fmt
            if c_rev30 is not None:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=c_rev30).number_format = rub_fmt
            for c in [c_cogs, c_fee, c_log, c_acq, c_total, c_buy_total, c_profit_unit, c_profit_batch]:
                if c is not None:
                    for row in range(2, worksheet.max_row + 1):
                        worksheet.cell(row=row, column=c).number_format = rub_fmt
            for c in [c_margin, c_roi]:
                if c is not None:
                    for row in range(2, worksheet.max_row + 1):
                        worksheet.cell(row=row, column=c).number_format = pct_fmt

        _apply_header_style(worksheet)
        # Зелёные колонки, которые заполняет пользователь
        _apply_user_input_column_style(worksheet, ["Кол-во к закупке", "Себестоимость"])
        _apply_fixed_column_widths_like_example(worksheet)

    out.seek(0)
    return out


def create_category_template(categories):
    """Создает Excel шаблон со ВСЕМИ категориями"""
    if not categories:
        return None

    data = []
    for i, cat in enumerate(categories, 1):
        path = cat.get('path', '')
        name = cat.get('name', '')

        # Разделяем путь на основную категорию и подкатегории
        path_parts = path.split('/') if path else []

        main_category = path_parts[0] if len(path_parts) > 0 else name
        subcategory = '/'.join(path_parts[1:]) if len(path_parts) > 1 else ''

        data.append({
            '№': i,
            'Категория': name,
            'Основная категория': main_category,
            'Подкатегория': subcategory,
            'Полный путь': path,
            'Выбрать': 'НЕТ'
        })

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Категории')

        worksheet = writer.sheets['Категории']
        worksheet.column_dimensions['A'].width = 8  # №
        worksheet.column_dimensions['B'].width = 50  # Категория
        worksheet.column_dimensions['C'].width = 30  # Основная категория
        worksheet.column_dimensions['D'].width = 50  # Подкатегория
        worksheet.column_dimensions['E'].width = 80  # Полный путь
        worksheet.column_dimensions['F'].width = 10  # Выбрать

        worksheet['G1'] = '📌 ИНСТРУКЦИЯ:'
        worksheet['G2'] = '1. Всего категорий: ' + str(len(categories))
        worksheet['G3'] = '2. Поставьте "ДА" в колонке "Выбрать" для нужных категорий'
        worksheet['G4'] = '3. Можно фильтровать по основной категории'
        worksheet['G5'] = '4. Сохраните файл и загрузите обратно'

    output.seek(0)
    return output


def parse_categories_from_excel(file_bytes, apply_exclusions=False):
    """Парсит загруженный Excel файл и возвращает список выбранных категорий
    apply_exclusions - если True, применяет список исключений (для интерфейса)
    """
    try:
        xls = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)

        def norm_col(c) -> str:
            return str(c).strip().lower()

        def pick_sheet():
            for _, df in xls.items():
                if df is None or df.empty:
                    continue
                cols_norm = {norm_col(c): c for c in df.columns}
                has_full_path = "полный путь" in cols_norm
                has_pair = ("категория" in cols_norm and "путь" in cols_norm)
                if has_full_path or has_pair:
                    return df, cols_norm
            return None, None

        df, cols_norm = pick_sheet()
        if df is None:
            return None

        selected = []

        choose_column = cols_norm.get("выбрать")

        if choose_column:
            for _, row in df.iterrows():
                choose_value = str(row.get(choose_column, "")).strip().lower()
                if choose_value in {"да", "yes", "1", "true", "y"}:
                    # Определяем, откуда брать путь
                    full_path_col = cols_norm.get("полный путь")
                    if full_path_col is not None:
                        path = row.get(full_path_col)
                        name = path.split('/')[-1] if path else ''
                    else:
                        name = row.get(cols_norm.get("категория"))
                        path = row.get(cols_norm.get("путь"))

                    # Применяем исключения только если apply_exclusions=True
                    if apply_exclusions:
                        from categories import is_allowed_category
                        if not is_allowed_category(name, path):
                            continue

                    selected.append({
                        'name': str(name or ""),
                        'path': str(path or ""),
                        'user_defined': True
                    })
        else:
            # Если нет колонки выбора, берем все строки
            for _, row in df.iterrows():
                full_path_col = cols_norm.get("полный путь")
                if full_path_col is not None:
                    path = row.get(full_path_col)
                    name = path.split('/')[-1] if path else ''
                else:
                    name = row.get(cols_norm.get("категория"))
                    path = row.get(cols_norm.get("путь"))

                # Применяем исключения только если apply_exclusions=True
                if apply_exclusions:
                    from categories import is_allowed_category
                    if not is_allowed_category(name, path):
                        continue

                selected.append({
                    'name': str(name or ""),
                    'path': str(path or ""),
                    'user_defined': True
                })

        return selected

    except Exception as e:
        print(f"Ошибка парсинга Excel: {e}")
        return None