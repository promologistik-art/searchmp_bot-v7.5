import io
import pandas as pd
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict

logger = logging.getLogger(__name__)


def _apply_trend_colors(worksheet, trend_column_number: int, max_row: int):
    """Применяет цветовую заливку для ячеек с трендом"""
    up_fill = PatternFill("solid", fgColor="FFC6EFCE")
    down_fill = PatternFill("solid", fgColor="FFFFC7CE")
    stable_fill = PatternFill("solid", fgColor="FFFFEB9C")
    na_fill = PatternFill("solid", fgColor="FFE0E0E0")
    
    for row in range(4, max_row + 1):
        cell = worksheet.cell(row=row, column=trend_column_number)
        value = str(cell.value).lower() if cell.value else ""
        
        if value in ["восходящий", "вверх"]:
            cell.fill = up_fill
            cell.value = "↑ вверх"
        elif value in ["нисходящий", "вниз"]:
            cell.fill = down_fill
            cell.value = "↓ вниз"
        elif value in ["стабильный", "ровно"]:
            cell.fill = stable_fill
            cell.value = "→ ровно"
        else:
            cell.fill = na_fill
            cell.value = "? нет данных"


def _apply_fixed_column_widths(worksheet):
    """Фиксированные ширины столбцов"""
    widths = {
        "A": 17.0,   # Ссылка на Ozon
        "B": 17.0,   # Категория
        "C": 17.0,   # Название товара
        "D": 10.0,   # Цена, р
        "E": 10.0,   # Кол-во продаж
        "F": 13.5,   # Выручка за 30 дней
        "G": 12.0,   # Кол-во конкурентов
        "H": 10.0,   # Тренд (3 мес)
        "I": 10.0,   # Кол-во к закупу
        "J": 10.0,   # Себестоимость
        "K": 10.0,   # % Комиссии
        "L": 10.0,   # Комиссия
        "M": 10.0,   # Логистика
        "N": 11.0,   # Эквайринг
        "O": 10.0,   # Налоги
        "P": 11.0,   # Всего расходы на единицу
        "Q": 11.0,   # Закуп итого
        "R": 11.0,   # Прибыль на ед
        "S": 11.0,   # Прибыль на партию
        "T": 11.0,   # Прибыль после налогов на ед
        "U": 9.0,   # Маржа, %
        "V": 9.0,   # ROI, %
    }
    for col, width in widths.items():
        try:
            worksheet.column_dimensions[col].width = width
            worksheet.column_dimensions[col].auto_size = False
        except:
            pass


def _apply_header_style(worksheet):
    """Стиль заголовков (строка 3)"""
    fill = PatternFill("solid", fgColor="FFEFEFEF")
    font = Font(bold=True)
    alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=3, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def _apply_green_cells(worksheet, max_row: int):
    """Зеленые ячейки для ввода пользователя"""
    green_fill = PatternFill("solid", fgColor="FFC6EFCE")
    
    # Зеленая ячейка B2
    worksheet.cell(row=2, column=2).fill = green_fill
    worksheet.cell(row=2, column=2).number_format = '0%'
    
    # Зеленые колонки: I (Кол-во к закупу), J (Себестоимость)
    for col in [9, 10]:  # I и J
        for row in range(4, max_row + 1):
            worksheet.cell(row=row, column=col).fill = green_fill


def _add_top_rows(worksheet):
    """Добавляет строки 1-2 как в образце"""
    # Строка 1: желтая инструкция
    worksheet.cell(row=1, column=1, value="Редактируйте только зеленые ячейки")
    worksheet.cell(row=1, column=1).font = Font(bold=True)
    
    # Строка 2: налог
    worksheet.cell(row=2, column=1, value="Налог")
    worksheet.cell(row=2, column=2, value=0.06)
    worksheet.cell(row=2, column=2).number_format = '0%'
    return "B2"


def create_excel_report(results: List[Dict]) -> io.BytesIO:
    """Создает Excel отчет с результатами анализа"""
    if not results:
        df = pd.DataFrame([{'Статус': 'Нет данных'}])
    else:
        df = pd.DataFrame(results)

        # Переименование колонок
        df = df.rename(columns={
            "category": "Категория",
            "name": "Название товара",
            "price": "Цена, р",
            "revenue": "Выручка за 30 дней",
            "competitors": "Кол-во конкурентов",
            "trend": "Тренд (3 мес)",
            "sales": "Кол-во продаж",
        })

        # Удаляем лишние колонки
        for drop_col in ("brand", "seller"):
            if drop_col in df.columns:
                df = df.drop(columns=[drop_col])

        if "url" in df.columns:
            df["Ссылка на Ozon"] = df["url"]
            df = df.drop(columns=["url"])

        # Пользовательские/расчётные колонки
        df["Кол-во к закупу"] = ""
        df["Себестоимость"] = ""
        df["% Комиссии"] = df.get("commission_percent", "")/100
        df["Комиссия"] = df.get("commission", 0)
        df["Логистика"] = df.get("logistics", 0)
        df["Эквайринг"] = ""
        df["Налоги"] = ""
        df["Всего расходы на единицу"] = ""
        df["Закуп итого"] = ""
        df["Прибыль на ед"] = ""
        df["Прибыль на партию"] = ""
        df["Прибыль после налогов"] = ""
        df["Маржа, %"] = ""
        df["ROI, %"] = ""

        # Удаляем временные колонки
        for col in ["commission", "commission_percent", "logistics"]:
            if col in df.columns:
                df = df.drop(columns=[col])

        # Порядок колонок как в образце
        col_order = [
            "Ссылка на Ozon",
            "Категория",
            "Название товара",
            "Цена, р",
            "Кол-во продаж",
            "Выручка за 30 дней",
            "Кол-во конкурентов",
            "Тренд (3 мес)",
            "Кол-во к закупу",
            "Себестоимость",
            "% Комиссии",
            "Комиссия",
            "Логистика",
            "Эквайринг",
            "Налоги",
            "Всего расходы на единицу",
            "Закуп итого",
            "Прибыль на ед",
            "Прибыль на партию",
            "Прибыль после налогов",
            "Маржа, %",
            "ROI, %",
        ]
        df = df[[c for c in col_order if c in df.columns]]

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Результаты анализа')
        worksheet = writer.sheets['Результаты анализа']

        # Вставляем 2 строки сверху (для инструкции и налога)
        worksheet.insert_rows(0, amount=2)
        
        # Добавляем верхние строки
        tax_cell_ref = _add_top_rows(worksheet)
        
        # Заголовки на строку 3
        for col_idx, cell in enumerate(df.columns, 1):
            worksheet.cell(row=3, column=col_idx, value=cell)
        
        # Данные с 4 строки
        for row_idx, row_data in enumerate(df.values, 4):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Получаем индексы колонок
        headers = [worksheet.cell(3, c).value for c in range(1, worksheet.max_column + 1)]
        
        def col_idx(name):
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        c_link = col_idx("Ссылка на Ozon")
        c_price = col_idx("Цена, р")
        c_sales_qty = col_idx("Кол-во продаж")
        c_qty = col_idx("Кол-во к закупу")
        c_cogs = col_idx("Себестоимость")
        c_comm_percent = col_idx("% Комиссии")
        c_comm_rub = col_idx("Комиссия")
        c_log = col_idx("Логистика")
        c_acq = col_idx("Эквайринг")
        c_tax = col_idx("Налоги")
        c_total = col_idx("Всего расходы на единицу")
        c_buy_total = col_idx("Закуп итого")
        c_profit_unit = col_idx("Прибыль на ед")
        c_profit_batch = col_idx("Прибыль на партию")
        c_profit_after_tax = col_idx("Прибыль после налогов")
        c_margin = col_idx("Маржа, %")
        c_roi = col_idx("ROI, %")
        c_rev30 = col_idx("Выручка за 30 дней")
        c_trend = col_idx("Тренд (3 мес)")

        max_row = worksheet.max_row

        for row in range(4, max_row + 1):
            # Ссылка
            if c_link:
                cell = worksheet.cell(row=row, column=c_link)
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    url = cell.value.replace('"', '""')
                    cell.value = f'=HYPERLINK("{url}","{url}")'
                    cell.style = "Hyperlink"

            # Эквайринг = Цена * 1.5%
            if c_acq and c_price:
                worksheet.cell(row=row, column=c_acq, value=f"={get_column_letter(c_price)}{row}*0.015")

            # Налоги = Цена * Налоговая ставка (как в образце: =D4*B2)
            if c_tax and c_price:
                worksheet.cell(row=row, column=c_tax, value=f"={get_column_letter(c_price)}{row}*{tax_cell_ref}")

            # Всего расходы на единицу = Себестоимость + Комиссия + Логистика + Эквайринг + Налоги
            expense_cols = []
            if c_cogs:
                expense_cols.append(get_column_letter(c_cogs))
            if c_comm_rub:
                expense_cols.append(get_column_letter(c_comm_rub))
            if c_log:
                expense_cols.append(get_column_letter(c_log))
            if c_acq:
                expense_cols.append(get_column_letter(c_acq))
            if c_tax:
                expense_cols.append(get_column_letter(c_tax))
            if c_total and expense_cols:
                worksheet.cell(row=row, column=c_total, value=f"={'+'.join([f'{col}{row}' for col in expense_cols])}")

            # Закуп итого = Кол-во к закупу * Себестоимость
            if c_buy_total and c_qty and c_cogs:
                worksheet.cell(row=row, column=c_buy_total, value=f"={get_column_letter(c_qty)}{row}*{get_column_letter(c_cogs)}{row}")

            # Прибыль на ед = Цена - Всего расходы на единицу
            if c_profit_unit and c_price and c_total:
                worksheet.cell(row=row, column=c_profit_unit, value=f"={get_column_letter(c_price)}{row}-{get_column_letter(c_total)}{row}")

            # Прибыль на партию = Прибыль на ед * Кол-во к закупу
            if c_profit_batch and c_profit_unit and c_qty:
                worksheet.cell(row=row, column=c_profit_batch, value=f"={get_column_letter(c_profit_unit)}{row}*{get_column_letter(c_qty)}{row}")

            # Прибыль после налогов = Прибыль на ед - Налоги (в образце =R4-O4)
            if c_profit_after_tax and c_profit_unit and c_tax:
                worksheet.cell(row=row, column=c_profit_after_tax, value=f"={get_column_letter(c_profit_unit)}{row}-{get_column_letter(c_tax)}{row}")

            # Маржа = Прибыль после налогов / Цена * 100 (как в образце: =IF(D4>0,T4/D4*100,""))
            if c_margin and c_profit_after_tax and c_price:
                worksheet.cell(row=row, column=c_margin, value=f"=IF({get_column_letter(c_price)}{row}>0,{get_column_letter(c_profit_after_tax)}{row}/{get_column_letter(c_price)}{row},\"\")")

            # ROI = Прибыль после налогов / Всего расходы * 100 (как в образце: =IF(P4>0,T4/P4*100,""))
            if c_roi and c_profit_after_tax and c_total:
                worksheet.cell(row=row, column=c_roi, value=f"=IF({get_column_letter(c_total)}{row}>0,{get_column_letter(c_profit_after_tax)}{row}/{get_column_letter(c_total)}{row},\"\")")

        # Цвета тренда
        if c_trend:
            _apply_trend_colors(worksheet, c_trend, max_row)

        # Форматы чисел
        rub_fmt = '#,##0\\ _₽'
        for row in range(4, max_row + 1):
            # Цена, Выручка, Себестоимость, Комиссия, Логистика, Эквайринг, Налоги, Расходы, Закуп, Прибыли
            for c in [c_price, c_rev30, c_cogs, c_comm_rub, c_log, c_acq, c_tax, c_total, c_buy_total, c_profit_unit, c_profit_batch, c_profit_after_tax]:
                if c:
                    worksheet.cell(row=row, column=c).number_format = rub_fmt
            # Проценты
            if c_comm_percent:
                worksheet.cell(row=row, column=c_comm_percent).number_format = '0%'
            if c_margin:
                worksheet.cell(row=row, column=c_margin).number_format = '0%'
            if c_roi:
                worksheet.cell(row=row, column=c_roi).number_format = '0%'
            # Целые числа для продаж
            if c_sales_qty:
                worksheet.cell(row=row, column=c_sales_qty).number_format = '#,##0'

        # Стили
        _apply_header_style(worksheet)
        _apply_green_cells(worksheet, max_row)
        _apply_fixed_column_widths(worksheet)

    out.seek(0)
    return out


def create_category_template(categories):
    """Создает Excel шаблон со ВСЕМИ категориями"""
    if not categories:
        return None

    data = []
    for i, cat in enumerate(categories, 1):
        path = cat.get('path', '')
        name = cat.get('name', '')

        path_parts = path.split('/') if path else []

        main_category = path_parts[0] if len(path_parts) > 0 else name
        subcategory = '/'.join(path_parts[1:]) if len(path_parts) > 1 else ''

        data.append({
            '№': i,
            'Категория': name,
            'Основная категория': main_category,
            'Подкатегория': subcategory,
            'Полный путь': path,
            'Выбрать': 'НЕТ'
        })

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Категории')

        worksheet = writer.sheets['Категории']
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 50
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['D'].width = 50
        worksheet.column_dimensions['E'].width = 80
        worksheet.column_dimensions['F'].width = 10

        worksheet['G1'] = '📌 ИНСТРУКЦИЯ:'
        worksheet['G2'] = '1. Всего категорий: ' + str(len(categories))
        worksheet['G3'] = '2. Поставьте "ДА" в колонке "Выбрать" для нужных категорий'
        worksheet['G4'] = '3. Можно фильтровать по основной категории'
        worksheet['G5'] = '4. Сохраните файл и загрузите обратно'

    output.seek(0)
    return output


def parse_categories_from_excel(file_bytes, apply_exclusions=False):
    """Парсит загруженный Excel файл и возвращает список выбранных категорий"""
    try:
        xls = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)

        def norm_col(c) -> str:
            return str(c).strip().lower()

        def pick_sheet():
            for _, df in xls.items():
                if df is None or df.empty:
                    continue
                cols_norm = {norm_col(c): c for c in df.columns}
                has_full_path = "полный путь" in cols_norm
                has_pair = ("категория" in cols_norm and "путь" in cols_norm)
                if has_full_path or has_pair:
                    return df, cols_norm
            return None, None

        df, cols_norm = pick_sheet()
        if df is None:
            return None

        selected = []

        choose_column = cols_norm.get("выбрать")

        if choose_column:
            for _, row in df.iterrows():
                choose_value = str(row.get(choose_column, "")).strip().lower()
                if choose_value in {"да", "yes", "1", "true", "y"}:
                    full_path_col = cols_norm.get("полный путь")
                    if full_path_col is not None:
                        path = row.get(full_path_col)
                        name = path.split('/')[-1] if path else ''
                    else:
                        name = row.get(cols_norm.get("категория"))
                        path = row.get(cols_norm.get("путь"))

                    if apply_exclusions:
                        from categories import is_allowed_category
                        if not is_allowed_category(name, path):
                            continue

                    selected.append({
                        'name': str(name or ""),
                        'path': str(path or ""),
                        'user_defined': True
                    })
        else:
            for _, row in df.iterrows():
                full_path_col = cols_norm.get("полный путь")
                if full_path_col is not None:
                    path = row.get(full_path_col)
                    name = path.split('/')[-1] if path else ''
                else:
                    name = row.get(cols_norm.get("категория"))
                    path = row.get(cols_norm.get("путь"))

                if apply_exclusions:
                    from categories import is_allowed_category
                    if not is_allowed_category(name, path):
                        continue

                selected.append({
                    'name': str(name or ""),
                    'path': str(path or ""),
                    'user_defined': True
                })

        return selected

    except Exception as e:
        print(f"Ошибка парсинга Excel: {e}")
        return None